# app/storage.py — 파일/시트 I/O + 변환 유틸
import os
from datetime import datetime, date
from openpyxl import Workbook, load_workbook
from .constants import (
    FILE_PATH, S_EX, S_LOG, S_PROF, S_MEAL,
    DEFAULT_DB, MEAL_HEADERS
)

try:
    from openpyxl.utils.datetime import from_excel as _ox_from_excel
except Exception:
    _ox_from_excel = None

# ------------- 초기화 -------------
def init_workbook():
    if os.path.exists(FILE_PATH):
        return
    wb = Workbook()

    ws_ex = wb.active
    ws_ex.title = S_EX
    ws_ex.append(["부위", "운동명"])
    for part, items in DEFAULT_DB.items():
        for ex in items:
            ws_ex.append([part, ex])

    ws_log = wb.create_sheet(S_LOG)
    ws_log.append(["날짜(YYYY-MM-DD)","닉네임","목표","분할","부위","운동명","세트수","반복수","무게(kg)","완료세트","총볼륨"])

    ws_pf = wb.create_sheet(S_PROF)
    ws_pf.append(["닉네임","성별","키(cm)","몸무게(kg)","나이","목표","업데이트시각"])

    ws_meal = wb.create_sheet(S_MEAL)
    ws_meal.append(["날짜(YYYY-MM-DD)","닉네임","식사분류","메뉴명","주성분","중량(g)","개수","횟수"])

    wb.save(FILE_PATH)

def load_wb():
    return load_workbook(FILE_PATH)

# ------------- 변환 유틸 -------------
def _to_date(cell_value):
    if cell_value in (None, ""):
        return None
    if isinstance(cell_value, date) and not isinstance(cell_value, datetime):
        return cell_value
    if isinstance(cell_value, datetime):
        return cell_value.date()
    if isinstance(cell_value, (int, float)) and _ox_from_excel:
        try:
            return _ox_from_excel(cell_value).date()
        except Exception:
            pass
    s = str(cell_value).strip()
    for fmt in ("%Y-%m-%d","%Y/%m/%d","%Y.%m.%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None

def _to_int(x, default=0):
    if x in (None, ""):
        return default
    try:
        return int(x)
    except ValueError:
        try:
            return int(float(x))
        except Exception:
            return default

# ------------- Exercises -------------
def read_exercises_by_part():
    wb = load_wb(); ws = wb[S_EX]
    data, header = {}, True
    for row in ws.iter_rows(values_only=True):
        if header: header=False; continue
        part, name = row
        if not part or not name: continue
        data.setdefault(part, [])
        if name not in data[part]: data[part].append(name)
    return data

# ------------- Logs -------------
def append_log_row(row):
    wb = load_wb(); ws = wb[S_LOG]
    ws.append(row); wb.save(FILE_PATH)

def replace_logs_for_date(nickname, d: date, new_items):
    wb = load_wb(); ws = wb[S_LOG]
    all_rows = list(ws.iter_rows(values_only=True))
    header, kept = all_rows[0], [all_rows[0]]

    for r in all_rows[1:]:
        d_str, nick, *_ = r
        try:
            rd = datetime.strptime(str(d_str), "%Y-%m-%d").date()
        except Exception:
            kept.append(r); continue
        if not (rd == d and nick == nickname):
            kept.append(r)

    ws.delete_rows(1, ws.max_row)
    for r in kept: ws.append(r)
    for it in (new_items or []):
        ws.append([
            d.strftime("%Y-%m-%d"), nickname, it["goal"], it["split"], it["part"], it["name"],
            int(it["sets"]), int(it["reps"]), int(it["weight"]), int(it.get("done_sets",0)),
            int(it.get("volume", int(it["weight"])*int(it["reps"])*int(it.get("done_sets",0))))
        ])
    wb.save(FILE_PATH)

def fetch_logs_month(nickname_filter, year, month):
    wb = load_wb(); ws = wb[S_LOG]
    out, header = [], True
    for row in ws.iter_rows(values_only=True):
        if header: header=False; continue
        d_str, nick, goal, split, part, ex_name, sets, reps, weight, done_sets, volume = row
        try:
            d = datetime.strptime(str(d_str), "%Y-%m-%d").date()
        except Exception:
            continue
        if d.year == year and d.month == month and (nickname_filter.strip()=="" or nick==nickname_filter):
            out.append({
                "date": d, "nick": nick, "goal": goal, "split": split, "part": part, "name": ex_name,
                "sets": int(sets), "reps": int(reps), "weight": int(weight),
                "done_sets": int(done_sets), "volume": int(volume)
            })
    return out

# ------------- Profiles -------------
def upsert_profile(nickname, gender, height, weight, age, goal):
    wb = load_wb()
    if S_PROF not in wb.sheetnames:
        ws_pf = wb.create_sheet(S_PROF)
        ws_pf.append(["닉네임","성별","키(cm)","몸무게(kg)","나이","목표","업데이트시각"])
    ws = wb[S_PROF]

    header, row_idx, idx = True, None, 0
    for row in ws.iter_rows(values_only=False):
        idx += 1
        if header: header=False; continue
        if row[0].value == nickname:
            row_idx = idx; break

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    if row_idx is None:
        ws.append([nickname, gender, height, weight, age, goal, now_str])
    else:
        ws.cell(row=row_idx, column=2).value = gender
        ws.cell(row=row_idx, column=3).value = height
        ws.cell(row=row_idx, column=4).value = weight
        ws.cell(row=row_idx, column=5).value = age
        ws.cell(row=row_idx, column=6).value = goal
        ws.cell(row=row_idx, column=7).value = now_str
    wb.save(FILE_PATH)

def load_profile(nickname):
    wb = load_wb()
    if S_PROF not in wb.sheetnames: return None
    ws, header = wb[S_PROF], True
    for row in ws.iter_rows(values_only=True):
        if header: header=False; continue
        nick, g, h, w, a, goal, _ = row
        if nick == nickname:
            return {"nickname": nick, "gender": g, "height": h, "weight": w, "age": a, "goal": goal}
    return None

# ------------- Meals -------------
def _ensure_meal_sheet(wb):
    if S_MEAL not in wb.sheetnames:
        ws = wb.create_sheet(S_MEAL); ws.append(MEAL_HEADERS); return ws, True
    ws = wb[S_MEAL]
    first = next(ws.iter_rows(values_only=True), None)
    if not first:
        ws.append(MEAL_HEADERS); return ws, True
    if all(v in (None,"") for v in first):
        ws.delete_rows(1, 1); ws.insert_rows(1)
        for c,h in enumerate(MEAL_HEADERS, start=1): ws.cell(row=1, column=c, value=h)
        return ws, True
    if list(first)[:len(MEAL_HEADERS)] != MEAL_HEADERS:
        ws.insert_rows(1)
        for c,h in enumerate(MEAL_HEADERS, start=1): ws.cell(row=1, column=c, value=h)
        return ws, True
    return ws, False

def append_meal_row(row):
    wb = load_wb(); ws, _ = _ensure_meal_sheet(wb)
    d = _to_date(row[0])
    if d is None: raise ValueError("append_meal_row: 'date'가 유효한 날짜가 아닙니다.")
    cleaned = [
        d.strftime("%Y-%m-%d"),
        (row[1] or "").strip(), (row[2] or "").strip(), (row[3] or "").strip(), (row[4] or "").strip(),
        _to_int(row[5]), _to_int(row[6]), _to_int(row[7]),
    ]
    ws.append(cleaned); wb.save(FILE_PATH)

def fetch_meals_by_date(nickname, d: date):
    wb = load_wb()
    if S_MEAL not in wb.sheetnames: return []
    ws, out = wb[S_MEAL], []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i == 1: continue
        if not row or all(v in (None,"") for v in row): continue
        d_str, nick, meal_type, menu, main, grams, count, times = (list(row)+[None]*8)[:8]
        rd = _to_date(d_str)
        if rd is None: continue
        nick_ok = True if (nickname is None or str(nickname).strip()=="") else (nick == nickname)
        if rd == d and nick_ok:
            out.append({
                "date": rd, "nick": nick, "meal_type": (meal_type or ""), "menu": (menu or ""),
                "main": (main or ""), "grams": _to_int(grams), "count": _to_int(count), "times": _to_int(times)
            })
    return out

def replace_meals_for_date(nickname, d: date, new_items):
    wb = load_wb(); ws, _ = _ensure_meal_sheet(wb)
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        ws.append(MEAL_HEADERS); all_rows = [tuple(MEAL_HEADERS)]
    header = list(all_rows[0]); kept = [header]
    for r in all_rows[1:]:
        if not r or all(v in (None,"") for v in r): continue
        d_str, nick, *_ = (list(r)+[None]*8)[:8]
        rd = _to_date(d_str)
        if rd is None or not (rd == d and nick == nickname): kept.append(r)
    ws.delete_rows(1, ws.max_row)
    for r in kept: ws.append(r)
    for it in (new_items or []):
        ws.append([
            d.strftime("%Y-%m-%d"), nickname,
            (it.get("meal_type") or "").strip(), (it.get("menu") or "").strip(),
            (it.get("main") or "").strip(), _to_int(it.get("grams",0)),
            _to_int(it.get("count",0)), _to_int(it.get("times",0)),
        ])
    wb.save(FILE_PATH)